import datetime
import requests
import openpyxl

# define a URL da API para obter a cotação do dólar
url = "https://economia.awesomeapi.com.br/json/USD-BRL"

# define o caminho do arquivo de Excel
arquivo_excel = r"C:\Users\masuc\OneDrive\Área de Trabalho\planilhas_financeiras\dolar.xlsx"


# define o nome da planilha e as células onde as cotações serão armazenadas
planilha = "dolar"
celula_dia = "A{}"
celula_mes = "B{}"
celula_ano = "C{}"

# define o índice inicial das cotações para cada período
indice_inicial_dia = 2
indice_inicial_mes = 2
indice_inicial_ano = 2

def atualizar_cotacao():
    # faz a requisição para a API
    resposta = requests.get(url)
    # obtém a data e o valor da cotação do dólar
    data = datetime.datetime.now()
    cotacao = float(resposta.json()[0]["bid"])
    # abre o arquivo de Excel
    workbook = openpyxl.load_workbook(arquivo_excel)
    # seleciona a planilha
    worksheet = workbook[planilha]
    # verifica se já existe uma cotação para o dia atual
    indice = indice_inicial_dia
    while worksheet[celula_dia.format(indice)].value is not None:
        data_celula = worksheet[celula_dia.format(indice)].value
        if data_celula.date() == data.date():
            break
        indice += 1
    else:
        # insere uma nova linha para a cotação do dia atual
        worksheet[celula_dia.format(indice)].value = data
        worksheet[celula_mes.format(indice)].value = data.strftime("%B")
        worksheet[celula_ano.format(indice)].value = data.year
    # insere a cotação do dólar na coluna correspondente
    if worksheet[celula_dia.format(indice)].value == data:
        worksheet.cell(row=indice, column=4).value = cotacao
    elif worksheet[celula_mes.format(indice)].value == data.strftime("%B"):
        worksheet.cell(row=indice, column=5).value = cotacao
    else:
        worksheet.cell(row=indice, column=6).value = cotacao
    # salva o arquivo de Excel
    workbook.save(arquivo_excel)
    # fecha o arquivo de Excel
    workbook.close()

# pede ao usuário para ativar a atualização
input("Pressione Enter para atualizar a cotação do dólar.")

# atualiza a cotação do dólar
atualizar_cotacao()
